from openpyxl import Workbook, load_workbook
import pandas
import operator

data = {}
full_buy = []
multiplier = [0, 0, 0.9, 0.5, 0.5, 1, 3, 0.45, 1, 0.5, 1, 1, 1, 0.3, 1, 2, 1, 1, 1, 2, 5, 1, 1, 1, 1]


def get_and_sort_data():
    file_read = load_workbook(filename='data-20200319T1305-structure-20200319T1305.xlsx', read_only=True)
    sheet_data = file_read['Лист 1 - data-20200319T1305-str']
    row = 3
    while row <= sheet_data.max_row:
        column = 4
        template_data = {}
        while column < sheet_data.max_column:
            try:
                template_data[str(sheet_data.cell(row=2, column=column).value).replace(', RUB.', '')] = float(sheet_data.cell(row=row, column=column).value)
            except Exception:
                template_data[str(sheet_data.cell(row=2, column=column).value).replace(', RUB.', '')] = 0
            finally:
                column += 1
        data[sheet_data["B{}".format(row)].value] = template_data
        row += 1
    the_cheapest_shop(sheet_data)
    avg_price(data)
    the_cheapest_buy(data)
    the_most_expensive_buy(data)


def the_cheapest_shop(xlsx_list):
    temp_top = {}
    column = 4
    while column < xlsx_list.max_column:
        mark = True
        row = 3
        summma = 0
        while row <= xlsx_list.max_row:
            try:
                summma += float(xlsx_list.cell(row=row, column=column).value)*multiplier[row-3]
            except Exception:
                mark = False
            finally:
                row += 1
        if mark:
            temp_top[str(xlsx_list.cell(row=2, column=column).value).replace(', RUB.', '')] = summma
        column += 1
    top = sorted(temp_top.items(), key=operator.itemgetter(1))
    top = dict(top)
    values = list(top.values())
    my_series = pandas.Series(values, index=top.keys(), name='Самые дешевые магазины для закупки на 1 человек на 1 неделю')
    print(my_series)


def avg_price(xlsx_list):
    avg_dict = {}
    for key in xlsx_list.keys():
        summa = 0
        shop_number = 0
        for value in xlsx_list.get(key).values():
            if value != 0:
                summa += float(value)
                shop_number += 1
        avg_dict[key] = summa/shop_number
    values = list(avg_dict.values())
    my_series = pandas.Series(values, index=avg_dict.keys(), name='Средная цена продуктов')
    print(my_series)


def the_cheapest_buy(xlsx_list):
    summa = 0
    for key in xlsx_list.keys():
        sorted_values = sorted(xlsx_list.get(key).items(), key=operator.itemgetter(1))
        summa += sorted_values[0][1]
    summa = pandas.Series([summa], index=['Минимальная цена за все продукты в одном магазине'])
    print(summa)


def the_most_expensive_buy(xlsx_list):
    summa = 0
    for key in xlsx_list.keys():
        sorted_values = sorted(xlsx_list.get(key).items(), key=operator.itemgetter(1))
        summa += sorted_values[-1][1]
    summa = pandas.Series([summa], index=['Максимальная цена за все продукты в одном магазине'])
    print(summa)


get_and_sort_data()
